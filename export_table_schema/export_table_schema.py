
import pymysql
import openpyxl
from openpyxl.styles import Side, Border, PatternFill, Font, Alignment


wb = openpyxl.Workbook()
sheet = wb.active

thin_border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
pattern_blue_color = PatternFill(start_color='00003366', end_color='00003366', fill_type='solid')
text_white_color = Font(color='00FFFFFF', bold=True)

HOST = input('INPUT HOST : ')
USER = input('INPUT USER : ')
PASSWD = input('INPUT PASSWD : ')
EXCEL_PATH = input('INPUT EXCEL PATH : ')

CONN = pymysql.connect(host=HOST, user=USER, passwd=PASSWD)
CURS = CONN.cursor()
TABLE_INFO = """
                SELECT  TABLE_NAME 
                	,	TABLE_COMMENT  
                	,	TABLE_COLLATION 
                FROM 	information_schema.TABLES 
                WHERE 	TABLE_SCHEMA = 'caredocplus'
                ORDER BY TABLE_NAME
                LIMIT 100 
"""
CURS.execute(TABLE_INFO) 
TABLE_RESULT = CURS.fetchall() 
CELL_NO = 0
print(len(TABLE_RESULT))

for TAB_NO in range(len(TABLE_RESULT)):


    TABLE_NAME = TABLE_RESULT[TAB_NO][0]
    TABLE_COMMENT = TABLE_RESULT[TAB_NO][1]
    TABLE_COLLATE = TABLE_RESULT[TAB_NO][2]


    CELL_NO += 1
    sheet.cell(row=CELL_NO, column=1, value="No." + str(TAB_NO+1)).border = thin_border
    sheet.cell(row=CELL_NO, column=1, value="No." + str(TAB_NO+1)).fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=1, value="No." + str(TAB_NO+1)).font = text_white_color
    sheet.cell(row=CELL_NO, column=1, value="No." + str(TAB_NO+1)).alignment = Alignment(horizontal='center',vertical='center')

    sheet.cell(row=CELL_NO, column=3, value="COMMENT").border = thin_border
    sheet.cell(row=CELL_NO, column=3, value="COMMENT").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=3, value="COMMENT").font = text_white_color

    sheet.cell(row=CELL_NO, column=5, value="CHARSET").border = thin_border
    sheet.cell(row=CELL_NO, column=5, value="CHARSET").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=5, value="CHARSET").font = text_white_color

    sheet.cell(row=CELL_NO, column=7, value="COLLATE").border = thin_border
    sheet.cell(row=CELL_NO, column=7, value="COLLATE").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=7, value="COLLATE").font = text_white_color


    sheet.cell(row=CELL_NO, column=2, value=TABLE_NAME).border = thin_border
    sheet.cell(row=CELL_NO, column=4, value=TABLE_COMMENT).border = thin_border
    sheet.cell(row=CELL_NO, column=6, value="utf8mb4").border = thin_border
    sheet.cell(row=CELL_NO, column=8, value=TABLE_COLLATE).border = thin_border
    

    CELL_NO += 1
    sheet.cell(row=CELL_NO, column=1, value="SEQ").border = thin_border
    sheet.cell(row=CELL_NO, column=1, value="SEQ").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=1, value="SEQ").font = text_white_color

    sheet.cell(row=CELL_NO, column=2, value="NAME").border = thin_border
    sheet.cell(row=CELL_NO, column=2, value="NAME").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=2, value="NAME").font = text_white_color

    sheet.cell(row=CELL_NO, column=3, value="TYPE").border = thin_border
    sheet.cell(row=CELL_NO, column=3, value="TYPE").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=3, value="TYPE").font = text_white_color

    sheet.cell(row=CELL_NO, column=4, value="NULLABLE").border = thin_border
    sheet.cell(row=CELL_NO, column=4, value="NULLABLE").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=4, value="NULLABLE").font = text_white_color

    sheet.cell(row=CELL_NO, column=5, value="DEFAULT").border = thin_border
    sheet.cell(row=CELL_NO, column=5, value="DEFAULT").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=5, value="DEFAULT").font = text_white_color

    sheet.cell(row=CELL_NO, column=6, value="EXTRA").border = thin_border
    sheet.cell(row=CELL_NO, column=6, value="EXTRA").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=6, value="EXTRA").font = text_white_color

    sheet.cell(row=CELL_NO, column=7, value="COLUMN COMMENT").border = thin_border
    sheet.cell(row=CELL_NO, column=7, value="COLUMN COMMENT").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=7, value="COLUMN COMMENT").font = text_white_color

    sheet.cell(row=CELL_NO, column=8, value="KEY").border = thin_border
    sheet.cell(row=CELL_NO, column=8, value="KEY").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=8, value="KEY").font = text_white_color



    COLUMN_INFO = """
                        SELECT  ORDINAL_POSITION
                      	    ,	COLUMN_NAME
                      	    ,   COLUMN_TYPE
                            , 	IS_NULLABLE
                            ,   COLUMN_DEFAULT
                      	    ,	REPLACE(EXTRA,"DEFAULT_GENERATED","")
                            ,   COLUMN_COMMENT
                            ,   COLUMN_KEY
                        FROM    information_schema.COLUMNS  
                        WHERE   TABLE_SCHEMA = 'caredocplus'
                        AND 	TABLE_NAME = '%s'
                        ORDER BY ORDINAL_POSITION;
    """ 
    CURS.execute(COLUMN_INFO % TABLE_NAME) 
    COLUMN_RESULT = CURS.fetchall() 

    for COL_NO in range(len(COLUMN_RESULT)):

        #print(COLUMN_RESULT[COL_NO])

        COLUMN_NO = COLUMN_RESULT[COL_NO][0]
        COLUMN_NAME = COLUMN_RESULT[COL_NO][1]
        COLUMN_TYPE = COLUMN_RESULT[COL_NO][2]

        COLUMN_NULLABLE = COLUMN_RESULT[COL_NO][3]
        if COLUMN_NULLABLE == "NO":
            COLUMN_NULLABLE = "NOT NULL"
        else:
            COLUMN_NULLABLE = "NULL"

        COLUMN_DEFAULT = COLUMN_RESULT[COL_NO][4]
        COLUMN_EXTRA = COLUMN_RESULT[COL_NO][5].strip()
        COLUMN_COMMENT = COLUMN_RESULT[COL_NO][6]
        COLUMN_KEY = COLUMN_RESULT[COL_NO][7]

        CELL_NO += 1
        sheet.cell(row=CELL_NO, column=1, value=COLUMN_NO).border = thin_border
        sheet.cell(row=CELL_NO, column=1, value=COLUMN_NO).alignment = Alignment(horizontal='center',vertical='center')

        sheet.cell(row=CELL_NO, column=2, value=COLUMN_NAME).border = thin_border
        sheet.cell(row=CELL_NO, column=3, value=COLUMN_TYPE).border = thin_border
        sheet.cell(row=CELL_NO, column=4, value=COLUMN_NULLABLE).border = thin_border
        sheet.cell(row=CELL_NO, column=5, value=COLUMN_DEFAULT).border = thin_border
        sheet.cell(row=CELL_NO, column=6, value=COLUMN_EXTRA).border = thin_border
        sheet.cell(row=CELL_NO, column=7, value=COLUMN_COMMENT).border = thin_border
        sheet.cell(row=CELL_NO, column=8, value=COLUMN_KEY).border = thin_border


    CELL_NO += 1
    sheet.merge_cells(start_row=CELL_NO, start_column=1, end_row=CELL_NO, end_column=8)
    sheet.cell(row=CELL_NO, column=1, value="INDEX").border = thin_border
    sheet.cell(row=CELL_NO, column=1, value="INDEX").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=1, value="INDEX").font = text_white_color


    CELL_NO += 1
    sheet.cell(row=CELL_NO, column=1, value="SEQ").border = thin_border
    sheet.cell(row=CELL_NO, column=1, value="SEQ").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=1, value="SEQ").font = text_white_color

    sheet.cell(row=CELL_NO, column=2, value="INDEX TYPE").border = thin_border
    sheet.cell(row=CELL_NO, column=2, value="INDEX TYPE").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=2, value="INDEX TYPE").font = text_white_color

    sheet.cell(row=CELL_NO, column=3, value="INDEX COLUMN").border = thin_border
    sheet.cell(row=CELL_NO, column=3, value="INDEX COLUMN").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=3, value="INDEX COLUMN").font = text_white_color

    sheet.cell(row=CELL_NO, column=4, value="INDEX NAME").border = thin_border
    sheet.cell(row=CELL_NO, column=4, value="INDEX NAME").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=4, value="INDEX NAME").font = text_white_color

    sheet.cell(row=CELL_NO, column=5, value="REFERENCE TABLE").border = thin_border
    sheet.cell(row=CELL_NO, column=5, value="REFERENCE TABLE").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=5, value="REFERENCE TABLE").font = text_white_color

    sheet.cell(row=CELL_NO, column=6, value="REFERENCE COLUMN").border = thin_border
    sheet.cell(row=CELL_NO, column=6, value="REFERENCE COLUMN").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=6, value="REFERENCE COLUMN").font = text_white_color

    sheet.cell(row=CELL_NO, column=7, value="DELETE RULE").border = thin_border
    sheet.cell(row=CELL_NO, column=7, value="DELETE RULE").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=7, value="DELETE RULE").font = text_white_color

    sheet.cell(row=CELL_NO, column=8, value="UPDATE RULE").border = thin_border
    sheet.cell(row=CELL_NO, column=8, value="UPDATE RULE").fill = pattern_blue_color
    sheet.cell(row=CELL_NO, column=8, value="UPDATE RULE").font = text_white_color


    INDEX_INFO = """
                        SELECT  a.SEQ_IN_INDEX
                        	, 	a.NON_UNIQUE
                        	,	a.COLUMN_NAME
                        	,	a.INDEX_NAME
                        	,	b.REFERENCED_TABLE_NAME
                        	,	b.REFERENCED_COLUMN_NAME
                        	,	b.DELETE_RULE 
                        	,	b.UPDATE_RULE 
                        FROM    information_schema.STATISTICS AS a 
                        LEFT OUTER JOIN (
                        					SELECT 	a.TABLE_NAME
                        					    ,	a.CONSTRAINT_NAME
                        						,	a.COLUMN_NAME
                        						,	a.REFERENCED_TABLE_NAME 
                                                ,	a.REFERENCED_COLUMN_NAME
                                                ,	b.DELETE_RULE
                                                ,	b.UPDATE_RULE 
                        					FROM 	information_schema.KEY_COLUMN_USAGE AS a
                        					INNER JOIN information_schema.REFERENTIAL_CONSTRAINTS AS b
                        					ON 	a.TABLE_SCHEMA = b.CONSTRAINT_SCHEMA AND a.CONSTRAINT_NAME = b.CONSTRAINT_NAME 
                        					WHERE   a.REFERENCED_TABLE_NAME IS NOT NULL
                        					AND 	a.TABLE_SCHEMA = 'caredocplus'
                        					AND 	a.TABLE_NAME = '%s'

                        ) AS b 
                        ON a.TABLE_NAME = b.TABLE_NAME AND a.COLUMN_NAME = b.COLUMN_NAME
                        WHERE   a.TABLE_SCHEMA = 'caredocplus'
                        AND 	a.TABLE_NAME = '%s'
                        ORDER BY a.SEQ_IN_INDEX, a.NON_UNIQUE;
    
    """
    CURS.execute(INDEX_INFO % (TABLE_NAME,TABLE_NAME)) 
    INDEX_RESULT = CURS.fetchall() 

    for IDX_NO in range(len(INDEX_RESULT)):

        print(INDEX_RESULT[IDX_NO])
        INDEX_NO =  INDEX_RESULT[IDX_NO][0]
        INDEX_UNIQUE_FLAG =  INDEX_RESULT[IDX_NO][1]
        COLUMN_NAME =  INDEX_RESULT[IDX_NO][2]
        INDEX_NAME = INDEX_RESULT[IDX_NO][3].strip()
        REFERENCED_TABLE_NAME = INDEX_RESULT[IDX_NO][4]
        REFERENCED_COLUMN_NAME = INDEX_RESULT[IDX_NO][5]
        DELETE_RULE = INDEX_RESULT[IDX_NO][6]
        UPDATE_RULE = INDEX_RESULT[IDX_NO][7]

        if INDEX_UNIQUE_FLAG == 0 and INDEX_NAME == "PRIMARY" and REFERENCED_TABLE_NAME is None:
            INDEX_UNIQUE_FLAG = "PK"
        elif INDEX_UNIQUE_FLAG == 0 and INDEX_NAME == "PRIMARY" and REFERENCED_TABLE_NAME is not None:
            INDEX_UNIQUE_FLAG = "PK, FK"
        elif INDEX_UNIQUE_FLAG == 0 and INDEX_NAME != "PRIMARY" and REFERENCED_TABLE_NAME is None:
            INDEX_UNIQUE_FLAG = "UK"                
        elif INDEX_UNIQUE_FLAG == 0 and INDEX_NAME != "PRIMARY" and REFERENCED_TABLE_NAME is not None:
            INDEX_UNIQUE_FLAG = "UK, FK"
        elif INDEX_UNIQUE_FLAG == 1 and REFERENCED_TABLE_NAME is None:
            INDEX_UNIQUE_FLAG = "INDEX"            
        elif INDEX_UNIQUE_FLAG == 1 and REFERENCED_TABLE_NAME is not None:
            INDEX_UNIQUE_FLAG = "FK"                                        
        
        CELL_NO += 1
  
        sheet.cell(row=CELL_NO, column=1, value=INDEX_NO).border = thin_border
        sheet.cell(row=CELL_NO, column=1, value=INDEX_NO).alignment = Alignment(horizontal='center',vertical='center')

        sheet.cell(row=CELL_NO, column=2, value=INDEX_UNIQUE_FLAG).border = thin_border
        sheet.cell(row=CELL_NO, column=3, value=COLUMN_NAME).border = thin_border
        sheet.cell(row=CELL_NO, column=4, value=INDEX_NAME).border = thin_border
        sheet.cell(row=CELL_NO, column=5, value=REFERENCED_TABLE_NAME).border = thin_border
        sheet.cell(row=CELL_NO, column=6, value=REFERENCED_COLUMN_NAME).border = thin_border
        sheet.cell(row=CELL_NO, column=7, value=DELETE_RULE).border = thin_border
        sheet.cell(row=CELL_NO, column=8, value=UPDATE_RULE).border = thin_border

    CELL_NO += 1
    sheet.merge_cells(start_row=CELL_NO, start_column=1, end_row=CELL_NO+1, end_column=8)       
    CELL_NO += 1

CURS.close()
CONN.close()

sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['B'].width = 50
sheet.column_dimensions['C'].width = 35
sheet.column_dimensions['D'].width = 45
sheet.column_dimensions['E'].width = 35
sheet.column_dimensions['F'].width = 35
sheet.column_dimensions['G'].width = 35
sheet.column_dimensions['H'].width = 25

EXCEL_PATH
wb.save(filename=EXCEL_PATH + "/test_info.xlsx")
